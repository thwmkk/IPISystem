"""Генерация отчёта в формате Excel (openpyxl)."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Стили
HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1E3A5F')
SECTION_FONT = Font(name='Calibri', size=14, bold=True, color='1E3A5F')
LABEL_FONT = Font(name='Calibri', size=11, bold=True, color='1A1A1A')
BODY_FONT = Font(name='Calibri', size=11, color='1A1A1A')
SMALL_FONT = Font(name='Calibri', size=10, color='555555')
ROW_ALT_FILL = PatternFill('solid', fgColor='F5F8FC')

THIN_BORDER = Border(
    left=Side(border_style='thin', color='CCCCCC'),
    right=Side(border_style='thin', color='CCCCCC'),
    top=Side(border_style='thin', color='CCCCCC'),
    bottom=Side(border_style='thin', color='CCCCCC'),
)


def _safe_title(title, used):
    r"""Безопасное имя листа: <=31 символ, без : \ / ? * [ ]."""
    bad = ':\\/?*[]'
    for c in bad:
        title = title.replace(c, '_')
    title = title[:30]
    base = title
    i = 2
    while title in used:
        suffix = f' ({i})'
        title = base[:30 - len(suffix)] + suffix
        i += 1
    used.add(title)
    return title


def _set_cell(ws, row, col, value, font=BODY_FONT, fill=None, border=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    if align is not None:
        cell.alignment = align
    return cell


def _write_section_header(ws, row, text, span=5):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _write_table(ws, start_row, headers, rows, col_widths=None):
    """Записать таблицу с заголовком и данными. Возвращает следующий row."""
    # Заголовок
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Данные
    for ri, row_data in enumerate(rows, start=1):
        actual_row = start_row + ri
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=actual_row, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if ri % 2 == 0:
                cell.fill = ROW_ALT_FILL
    # Ширины колонок
    if col_widths:
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    return start_row + len(rows) + 2  # пропуск 1 пустой строки


def _write_employee_sheet(wb, emp_data, used_titles):
    """Создаёт лист для одного сотрудника."""
    title = _safe_title(emp_data['employee']['full_name'], used_titles)
    ws = wb.create_sheet(title=title)

    row = 1
    # Заголовок
    cell = ws.cell(row=row, column=1, value=emp_data['employee']['full_name'])
    cell.font = Font(name='Calibri', size=18, bold=True, color='1E3A5F')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Профиль
    row = _write_section_header(ws, row, 'Профиль сотрудника')
    profile = emp_data['employee']
    profile_rows = [
        ('Должность:', profile['position']),
        ('Подразделение:', profile['department']),
        ('Возраст:', profile['age']),
        ('Стаж научной работы:', f"{profile['experience']} лет"),
        ('Учёная степень:', profile['academic_degree']),
        ('Email:', profile['email']),
    ]
    for k, v in profile_rows:
        _set_cell(ws, row, 1, k, font=LABEL_FONT)
        _set_cell(ws, row, 2, v, font=BODY_FONT)
        row += 1
    row += 1

    # IPI Сводка
    row = _write_section_header(ws, row, 'Индивидуальный показатель деятельности (IPI)')
    ipi = emp_data['ipi']
    ipi_rows = [
        ('Итоговый IPI:', round(ipi['total'], 2)),
        ('Научная составляющая:', round(ipi['scientific'], 2)),
        ('Организационная составляющая:', round(ipi['organizational'], 2)),
        ('Техническая составляющая:', round(ipi['technical'], 2)),
        ('Коэффициент возраста:', ipi['factors']['k_age']),
        ('Коэффициент стажа:', ipi['factors']['k_exp']),
        ('Коэффициент аспирантуры:', ipi['factors']['k_phd']),
    ]
    for k, v in ipi_rows:
        _set_cell(ws, row, 1, k, font=LABEL_FONT)
        _set_cell(ws, row, 2, v, font=BODY_FONT)
        row += 1
    row += 1

    # Научные работы
    row = _write_section_header(ws, row, f"Научные работы ({len(emp_data['scientific_works'])} шт.)")
    if emp_data['scientific_works']:
        rows = [
            (w['title'], w['work_type'], w['date'], w['points'],
             'Да' if w['verified'] else 'Нет')
            for w in emp_data['scientific_works']
        ]
        row = _write_table(ws, row,
                           ['Название', 'Тип', 'Дата', 'Баллы', 'Верифицирована'],
                           rows, col_widths=[50, 25, 14, 10, 16])
    else:
        _set_cell(ws, row, 1, 'Работ за период нет', font=SMALL_FONT)
        row += 2

    # Организационные работы
    row = _write_section_header(ws, row, f"Организационные работы ({len(emp_data['organizational_works'])} шт.)")
    if emp_data['organizational_works']:
        rows = [
            (w['title'], w['work_type'], w['date'], w['points'],
             'Да' if w['verified'] else 'Нет')
            for w in emp_data['organizational_works']
        ]
        row = _write_table(ws, row,
                           ['Название', 'Тип', 'Дата', 'Баллы', 'Верифицирована'],
                           rows, col_widths=[50, 25, 14, 10, 16])
    else:
        _set_cell(ws, row, 1, 'Работ за период нет', font=SMALL_FONT)
        row += 2

    # Технические работы
    row = _write_section_header(ws, row, f"Технические работы ({len(emp_data['technical_works'])} шт.)")
    if emp_data['technical_works']:
        rows = [
            (w['title'], w['work_type'], w['date'], w['points'],
             'Да' if w['verified'] else 'Нет')
            for w in emp_data['technical_works']
        ]
        row = _write_table(ws, row,
                           ['Название', 'Тип', 'Дата', 'Баллы', 'Верифицирована'],
                           rows, col_widths=[50, 25, 14, 10, 16])
    else:
        _set_cell(ws, row, 1, 'Работ за период нет', font=SMALL_FONT)
        row += 2

    # Выполненные задачи
    row = _write_section_header(ws, row, f"Выполненные задачи ({len(emp_data['tasks_done'])} шт.)")
    if emp_data['tasks_done']:
        rows = [
            (t['title'], t['project'], t['priority'], t['deadline'], t['points'])
            for t in emp_data['tasks_done']
        ]
        row = _write_table(ws, row,
                           ['Название', 'Проект', 'Приоритет', 'Срок', 'Баллы'],
                           rows, col_widths=[50, 30, 14, 14, 10])
    else:
        _set_cell(ws, row, 1, 'Выполненных задач за период нет', font=SMALL_FONT)
        row += 2

    # Невыполненные / просроченные задачи
    row = _write_section_header(ws, row, f"Невыполненные задачи ({len(emp_data['tasks_pending'])} шт.)")
    if emp_data['tasks_pending']:
        rows = [
            (t['title'], t['project'], t['priority'], t['status'], t['deadline'])
            for t in emp_data['tasks_pending']
        ]
        row = _write_table(ws, row,
                           ['Название', 'Проект', 'Приоритет', 'Статус', 'Срок'],
                           rows, col_widths=[50, 30, 14, 14, 14])
    else:
        _set_cell(ws, row, 1, 'Невыполненных задач нет', font=SMALL_FONT)
        row += 2

    # Проекты
    row = _write_section_header(ws, row, f"Проекты ({len(emp_data['projects'])} шт.)")
    if emp_data['projects']:
        rows = [
            (p['name'],
             'Создатель' if p['is_creator'] else 'Участник',
             p['start_date'],
             p['end_date'] or '—',
             'Завершён' if p['completed'] else 'В работе')
            for p in emp_data['projects']
        ]
        row = _write_table(ws, row,
                           ['Название проекта', 'Роль', 'Начало', 'Окончание', 'Статус'],
                           rows, col_widths=[50, 18, 14, 14, 14])
    else:
        _set_cell(ws, row, 1, 'Проектов нет', font=SMALL_FONT)
        row += 2


def render_department_excel(data) -> bytes:
    """Сгенерировать Excel-отчёт по отделу."""
    wb = Workbook()
    # Лист «Сводка»
    ws = wb.active
    ws.title = 'Сводка'
    used_titles = {'Сводка'}

    # Шапка
    cell = ws.cell(row=1, column=1,
                   value=f"Отчёт по деятельности отдела «{data['department']['short_name']}»")
    cell.font = Font(name='Calibri', size=16, bold=True, color='1E3A5F')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value=f"Период: {data['period']['label']}").font = BODY_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=3, column=1, value=f"Сформировано: {data['generated_at']}").font = SMALL_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

    # Таблица сводная
    headers = ['ФИО', 'Должность', 'IPI', 'Научная', 'Организационная', 'Техническая']
    rows = []
    for emp_data in data['employees']:
        rows.append((
            emp_data['employee']['full_name'],
            emp_data['employee']['position'],
            round(emp_data['ipi']['total'], 2),
            round(emp_data['ipi']['scientific'], 2),
            round(emp_data['ipi']['organizational'], 2),
            round(emp_data['ipi']['technical'], 2),
        ))
    _write_table(ws, 5, headers, rows,
                 col_widths=[35, 28, 12, 14, 18, 14])

    # По сотруднику отдельный лист
    for emp_data in data['employees']:
        _write_employee_sheet(wb, emp_data, used_titles)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_employee_excel(emp_data, department_short, period_label, generated_at) -> bytes:
    """Сгенерировать Excel-отчёт по одному сотруднику."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    # Шапка
    cell = ws.cell(row=1, column=1,
                   value=f"Отчёт по деятельности: {emp_data['employee']['full_name']}")
    cell.font = Font(name='Calibri', size=16, bold=True, color='1E3A5F')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    ws.cell(row=2, column=1, value=f"Отдел: {department_short}").font = BODY_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.cell(row=3, column=1, value=f"Период: {period_label}").font = BODY_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    ws.cell(row=4, column=1, value=f"Сформировано: {generated_at}").font = SMALL_FONT
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)

    _write_employee_sheet(wb, emp_data, used_titles={'Отчёт'})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
